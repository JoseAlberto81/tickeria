from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.utils import get_column_letter
from tickets.models import Ticket
from django.utils.timezone import localtime

@login_required
def exportar_excel_todos(request):
    from django.utils.timezone import localtime

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    # Encabezados
    headers = [
        "ID", "Estado", "Solicitante", "Correo electrónico", "Asignado",
        "Fecha esperada", "Prioridad", "Etiqueta", "Título", "Descripción", "Creado"
    ]
    ws.append(headers)

    # Datos
    for ticket in Ticket.objects.all().order_by('-creado'):
        ws.append([
            ticket.id,
            ticket.estado,
            ticket.solicitante,
            ticket.email,
            ticket.asignado,
            ticket.fecha_esperada.strftime('%d/%m/%Y') if ticket.fecha_esperada else "",
            ticket.prioridad,
            ticket.etiqueta,
            ticket.titulo,
            ticket.descripcion,
            localtime(ticket.creado).strftime('%d/%m/%Y %H:%M'),
        ])

    # Ajustar ancho de columnas
    for i, column in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    # Respuesta HTTP
    from django.http import HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="tickets.xlsx"'
    wb.save(response)
    return response

@login_required
def exportar_excel_mes(request):
    mes = request.GET.get('mes')  # formato: 'YYYY-MM'
    if not mes:
        return HttpResponse("Debes seleccionar un mes.", status=400)

    try:
        year, month = map(int, mes.split('-'))
    except Exception:
        return HttpResponse("Formato de mes inválido.", status=400)

    tickets = Ticket.objects.filter(
        creado__year=year,
        creado__month=month
    ).order_by('-creado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    headers = [
        "ID", "Estado", "Solicitante", "Correo electrónico", "Asignado",
        "Fecha esperada", "Prioridad", "Etiqueta", "Título", "Descripción", "Creado"
    ]
    ws.append(headers)

    for ticket in tickets:
        ws.append([
            ticket.id,
            ticket.estado,
            ticket.solicitante,
            ticket.email,
            ticket.asignado,
            ticket.fecha_esperada.strftime('%d/%m/%Y') if ticket.fecha_esperada else "",
            ticket.prioridad,
            ticket.etiqueta,
            ticket.titulo,
            ticket.descripcion,
            localtime(ticket.creado).strftime('%d/%m/%Y %H:%M'),
        ])

    # Ajustar ancho de columnas
    for i, column in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="tickets_{year}_{month:02d}.xlsx"'
    wb.save(response)
    return response